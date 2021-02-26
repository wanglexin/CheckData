import xlrd
import openpyxl
from tools import sqltools


class RWExcel(object):
    def readExcel(filename):
        wb = xlrd.open_workbook(filename=filename)
        sheet = wb.sheet_by_index(0)
        list = sheet.col_values(0)
        excelvalue = []
        for i in range(len(list)):
            if i > 0:
                excelvalue.append(int(list[i]))
            else:
                pass
        return excelvalue


    def SaveExcel(self, filename, x, y):
        wb = openpyxl.load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        ws.cell(row=x, column=5).value = y
        wb.save(filename)

    def SaveExcelforvalue(filename, x, y):
        wb = openpyxl.load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        ws.cell(row=x, column=7).value = y
        wb.save(filename)

    def readExcelToMoney(filename):
        wb = xlrd.open_workbook(filename=filename)
        sheet = wb.sheet_by_index(0)
        list = sheet.col_values(5)
        for i in range(len(list)):
            if list[i] != 0:
                lastmonth = sqltools.tools.lastmonth(sheet.cell_value(i, 0))
                nextmonth = sqltools.tools.nextmonth(sheet.cell_value(i, 0))
                total = sheet.cell_value(i, 3)
                a = (total + float(nextmonth[0][0]))-float(lastmonth[0][0])
                b = float(sheet.cell_value(i, 4))
                if a - b == 0:
                    x = "没有异常！"
                    RWExcel.SaveExcelforvalue(filename, i+1, x)
                else:
                    y = "不相等，数据异常！"
                    RWExcel.SaveExcelforvalue(filename, i+1, y)

            elif list[i] == 0:
                pass
            elif list[i] == "":
                pass



if __name__ == '__main__':
    filename = r"C:\Users\wanglexin\Desktop\对账.xlsx"
    a = RWExcel()
    a.readExcelToMoney(filename)
 #   a.SaveExcel(filename, 2, 1)